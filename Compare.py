from openpyxl import Workbook, load_workbook


class Compare:

    def __init__(self, fileBase, fileDst):
        print(fileBase)
        print(fileDst)
        self.wbA = load_workbook(fileBase.split('file:///')[1])
        self.wbB = load_workbook(fileDst.split('file:///')[1])
        self.fileDst = fileDst

    def getWbA(self):
        return self.wbA
    def getWbB(self):
        return self.wbB
    
    def getSheetsFromWb(self, wb):
        return wb.sheetnames()
    def compareExcel(self, sheetA, columnA, sheetB, columnB):
        colA = self.wbA[sheetA][columnA]
        colB = self.wbB[sheetB][columnB]
        self.wbB.create_sheet("Common A-B")
        self.wbB.create_sheet("No Common A-B")
        self.wbB.create_sheet("No Common B-A") 
        sheetCommonAB = self.wbB["Common A-B"]
        sheetNotCommonAB = self.wbB["No Common A-B"]
        sheetNotCommonBA = self.wbB["No Common B-A"]
        i = 1
        j = 1
        z = 1
        for rowA in colA:
            exist = False
            if rowA.value != None:
                for rowB in colB:
                    if str(rowA.value) == str(rowB.value):
                        exist = True
                        cellCommonsAB = 'A' + str(i)
                        sheetCommonAB[cellCommonsAB] = rowA.value
                        print(rowA.value)
                        i = i + 1
                if not exist:
                    cellNotCommonsAB = 'A' + str(j)
                    sheetNotCommonAB[cellNotCommonsAB] = rowA.value
                    j = j + 1
                exist = False
                   
        for rowB in colB:
            exist = False
            if rowB.value != None:
                for rowA in colA:
                    if str(rowB.value) == str(rowA.value):
                        exist = True
                if not exist:
                    cellNotCommonsBA = 'A' + str(z)
                    sheetNotCommonBA[cellNotCommonsBA] = rowB.value
                    z = z + 1
                exist = False
        self.wbB.save(( self.fileDst.split('file:///')[1] + ".result" ))
